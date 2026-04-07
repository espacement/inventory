import os
import sqlite3
import io
import json
import hashlib
import secrets
from datetime import datetime
from functools import wraps
from flask import (Flask, render_template, request, redirect,
                   url_for, send_file, flash, g, session)
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

DATABASE      = "inventory.db"
UPLOAD_FOLDER = "static/uploads"
ALLOWED_EXT   = {"png", "jpg", "jpeg", "webp", "gif"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

ITEM_STATUSES = {
    "active":   "✅ Активен",
    "broken":   "🔴 Сломан",
    "repair":   "🔧 В ремонте",
    "writeoff": "🗑 На списание",
    "reserved": "🔒 Зарезервирован",
}
MOVE_TYPES = {
    "receipt":  ("📥 Поступление", "#22c55e"),
    "expense":  ("📤 Расход",      "#f97316"),
    "writeoff": ("🗑 Списание",     "#ef4444"),
    "return":   ("↩️ Возврат",     "#a78bfa"),
}
UNITS = ["шт", "м", "кг", "л", "упак", "рулон", "пара", "компл"]

# ════════════════════════════════════════════════════════════
#  БД
# ════════════════════════════════════════════════════════════
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db: db.close()

def init_db():
    with app.app_context():
        db = get_db()
        db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT NOT NULL UNIQUE COLLATE NOCASE,
            password_hash TEXT NOT NULL,
            display_name  TEXT,
            role          TEXT DEFAULT 'user',
            created_at    TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL REFERENCES users(id),
            inv_num     TEXT NOT NULL,
            name        TEXT NOT NULL,
            location    TEXT,
            status      TEXT DEFAULT 'active',
            comment     TEXT,
            photo       TEXT,
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            updated_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE UNIQUE INDEX IF NOT EXISTS ux_item ON items(user_id, inv_num);
        CREATE TABLE IF NOT EXISTS nomenclature (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL REFERENCES users(id),
            code        TEXT,
            name        TEXT NOT NULL,
            unit        TEXT DEFAULT 'шт',
            quantity    REAL DEFAULT 0,
            min_qty     REAL DEFAULT 0,
            description TEXT,
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS movements (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL REFERENCES users(id),
            doc_num     TEXT NOT NULL,
            move_type   TEXT NOT NULL,
            nom_id      INTEGER NOT NULL REFERENCES nomenclature(id),
            quantity    REAL NOT NULL,
            comment     TEXT,
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE INDEX IF NOT EXISTS idx_items_u ON items(user_id);
        CREATE INDEX IF NOT EXISTS idx_nom_u   ON nomenclature(user_id);
        CREATE INDEX IF NOT EXISTS idx_mov_u   ON movements(user_id);
        CREATE INDEX IF NOT EXISTS idx_mov_n   ON movements(nom_id);
        """)
        db.commit()

def allowed_file(fn):
    return "." in fn and fn.rsplit(".", 1)[1].lower() in ALLOWED_EXT

# ════════════════════════════════════════════════════════════
#  AUTH
# ════════════════════════════════════════════════════════════
def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()

def current_user():
    uid = session.get("user_id")
    if not uid: return None
    return get_db().execute("SELECT * FROM users WHERE id=?", [uid]).fetchone()

def login_required(f):
    @wraps(f)
    def dec(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return dec

@app.context_processor
def inject_user(): return {"me": current_user()}

def uid(): return session["user_id"]

def next_doc(db, user_id, prefix="ДВ"):
    n = db.execute("SELECT COUNT(*) FROM movements WHERE user_id=?", [user_id]).fetchone()[0] + 1
    return f"{prefix}-{n:05d}"

def apply_mov(db, nom_id, move_type, qty):
    if move_type in ("receipt", "return"):
        db.execute("UPDATE nomenclature SET quantity=quantity+? WHERE id=?", [qty, nom_id])
    else:
        db.execute("UPDATE nomenclature SET quantity=MAX(0,quantity-?) WHERE id=?", [qty, nom_id])

# ─── Login ───────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    if session.get("user_id"): return redirect(url_for("index"))
    if request.method == "POST":
        u = request.form["username"].strip()
        p = request.form["password"]
        user = get_db().execute("SELECT * FROM users WHERE username=?", [u]).fetchone()
        if user and user["password_hash"] == hash_pw(p):
            session.permanent = True
            session["user_id"] = user["id"]
            flash(f"Добро пожаловать, {user['display_name'] or user['username']}!", "success")
            return redirect(url_for("index"))
        flash("Неверный логин или пароль.", "error")
    return render_template("auth.html", mode="login")

@app.route("/register", methods=["GET","POST"])
def register():
    if session.get("user_id"): return redirect(url_for("index"))
    if request.method == "POST":
        u    = request.form["username"].strip()
        name = request.form.get("display_name","").strip()
        p    = request.form["password"]
        p2   = request.form["password2"]
        err  = None
        if len(u) < 3:   err = "Логин минимум 3 символа."
        elif len(p) < 4: err = "Пароль минимум 4 символа."
        elif p != p2:    err = "Пароли не совпадают."
        else:
            db = get_db()
            if db.execute("SELECT id FROM users WHERE username=?", [u]).fetchone():
                err = "Логин занят."
        if err:
            flash(err, "error")
            return render_template("auth.html", mode="register")
        db = get_db()
        db.execute("INSERT INTO users(username,password_hash,display_name) VALUES(?,?,?)",
                   [u, hash_pw(p), name or u])
        db.commit()
        user = db.execute("SELECT * FROM users WHERE username=?", [u]).fetchone()
        session["user_id"] = user["id"]
        session.permanent  = True
        flash(f"Добро пожаловать, {name or u}!", "success")
        return redirect(url_for("index"))
    return render_template("auth.html", mode="register")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/profile", methods=["GET","POST"])
@login_required
def profile():
    db   = get_db()
    user = current_user()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "rename":
            dn = request.form.get("display_name","").strip()
            db.execute("UPDATE users SET display_name=? WHERE id=?", [dn, uid()])
            db.commit()
            flash("Имя обновлено.", "success")
        elif action == "password":
            old = request.form.get("old_password","")
            new = request.form.get("new_password","")
            n2  = request.form.get("new_password2","")
            if user["password_hash"] != hash_pw(old):  flash("Неверный текущий пароль.", "error")
            elif len(new) < 4:                          flash("Минимум 4 символа.", "error")
            elif new != n2:                             flash("Пароли не совпадают.", "error")
            else:
                db.execute("UPDATE users SET password_hash=? WHERE id=?", [hash_pw(new), uid()])
                db.commit()
                flash("Пароль изменён.", "success")
        return redirect(url_for("profile"))
    stats = {
        "items_count": db.execute("SELECT COUNT(*) FROM items WHERE user_id=?", [uid()]).fetchone()[0],
        "nom":        db.execute("SELECT COUNT(*) FROM nomenclature WHERE user_id=?", [uid()]).fetchone()[0],
        "mov":        db.execute("SELECT COUNT(*) FROM movements WHERE user_id=?", [uid()]).fetchone()[0],
    }
    return render_template("profile.html", user=user, stats=stats)

# ─── Экспорт / Импорт ────────────────────────────────────
@app.route("/export/json")
@login_required
def export_json():
    db = get_db()
    nom  = [dict(r) for r in db.execute("SELECT * FROM nomenclature WHERE user_id=?", [uid()])]
    itms = [dict(r) for r in db.execute("SELECT * FROM items WHERE user_id=?", [uid()])]
    movs = [dict(r) for r in db.execute("SELECT * FROM movements WHERE user_id=?", [uid()])]
    for lst in (nom, itms, movs):
        for row in lst: row.pop("user_id", None)
    data = {"exported_at": datetime.now().isoformat(), "version": "3",
            "nomenclature": nom, "items": itms, "movements": movs}
    out = io.BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode())
    return send_file(out, mimetype="application/json", as_attachment=True,
                     download_name=f"backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json")

@app.route("/import/json", methods=["GET","POST"])
@login_required
def import_json():
    if request.method == "POST":
        f = request.files.get("backup")
        if not f or not f.filename.endswith(".json"):
            flash("Выберите .json файл.", "error")
            return redirect(url_for("import_json"))
        try:
            data = json.loads(f.read().decode("utf-8"))
        except Exception:
            flash("Ошибка чтения файла.", "error")
            return redirect(url_for("import_json"))

        db   = get_db()
        mode = request.form.get("mode", "merge")
        if mode == "replace":
            db.execute("DELETE FROM movements WHERE user_id=?",   [uid()])
            db.execute("DELETE FROM nomenclature WHERE user_id=?", [uid()])
            db.execute("DELETE FROM items WHERE user_id=?",        [uid()])
            db.commit()

        nom_map, nc, ic, mc = {}, 0, 0, 0
        for n in data.get("nomenclature", []):
            old = n.get("id")
            try:
                cur = db.execute(
                    "INSERT INTO nomenclature(user_id,code,name,unit,quantity,min_qty,description,created_at)"
                    " VALUES(?,?,?,?,?,?,?,?)",
                    [uid(), n.get("code"), n["name"], n.get("unit","шт"),
                     n.get("quantity",0), n.get("min_qty",0),
                     n.get("description"), n.get("created_at")])
                nom_map[old] = cur.lastrowid; nc += 1
            except Exception: pass

        for it in data.get("items", []):
            try:
                db.execute(
                    "INSERT INTO items(user_id,inv_num,name,location,status,comment,created_at,updated_at)"
                    " VALUES(?,?,?,?,?,?,?,?)",
                    [uid(), it["inv_num"], it["name"], it.get("location"),
                     it.get("status","active"), it.get("comment"),
                     it.get("created_at"), it.get("updated_at")]); ic += 1
            except Exception: pass

        for m in data.get("movements", []):
            new_nom = nom_map.get(m.get("nom_id"))
            if not new_nom: continue
            try:
                db.execute(
                    "INSERT INTO movements(user_id,doc_num,move_type,nom_id,quantity,comment,created_at)"
                    " VALUES(?,?,?,?,?,?,?)",
                    [uid(), m["doc_num"], m["move_type"], new_nom,
                     m["quantity"], m.get("comment"), m.get("created_at")]); mc += 1
            except Exception: pass

        db.commit()
        flash(f"Импортировано: {nc} номенклатур, {ic} предметов, {mc} документов.", "success")
        return redirect(url_for("index"))
    return render_template("import.html")

# ════════════════════════════════════════════════════════════
#  ИМУЩЕСТВО
# ════════════════════════════════════════════════════════════
@app.route("/")
@login_required
def index():
    db = get_db()
    search = request.args.get("q","").strip()
    sf     = request.args.get("status","")
    q = "SELECT * FROM items WHERE user_id=?"
    p = [uid()]
    if search:
        q += " AND (inv_num LIKE ? OR name LIKE ? OR location LIKE ? OR comment LIKE ?)"
        like = f"%{search}%"; p += [like]*4
    if sf and sf in ITEM_STATUSES:
        q += " AND status=?"; p.append(sf)
    q += " ORDER BY updated_at DESC"
    items = db.execute(q, p).fetchall()
    return render_template("index.html", items=items,
                           statuses=ITEM_STATUSES, search=search, status_f=sf)

@app.route("/items/add", methods=["GET","POST"])
@app.route("/items/edit/<int:item_id>", methods=["GET","POST"])
@login_required
def item_form(item_id=None):
    db   = get_db()
    item = db.execute("SELECT * FROM items WHERE id=? AND user_id=?",
                      [item_id, uid()]).fetchone() if item_id else None
    if request.method == "POST":
        inv_num  = request.form["inv_num"].strip()
        name     = request.form["name"].strip()
        location = request.form.get("location","").strip()
        status   = request.form.get("status","active")
        comment  = request.form.get("comment","").strip()
        photo    = item["photo"] if item else None
        if not inv_num or not name:
            flash("Инвентарный номер и название обязательны!", "error")
            return render_template("item_form.html", statuses=ITEM_STATUSES, item=item)
        if db.execute("SELECT id FROM items WHERE inv_num=? AND user_id=? AND id!=?",
                      [inv_num, uid(), item_id or -1]).fetchone():
            flash(f"Номер «{inv_num}» уже существует!", "error")
            return render_template("item_form.html", statuses=ITEM_STATUSES, item=item)
        f = request.files.get("photo")
        if f and f.filename and allowed_file(f.filename):
            if photo:
                old = os.path.join(UPLOAD_FOLDER, photo)
                if os.path.exists(old): os.remove(old)
            fn = secure_filename(f"{uid()}_{inv_num}_{f.filename}")
            f.save(os.path.join(UPLOAD_FOLDER, fn)); photo = fn
        if item_id:
            db.execute("UPDATE items SET inv_num=?,name=?,location=?,status=?,comment=?,photo=?,"
                       "updated_at=datetime('now','localtime') WHERE id=? AND user_id=?",
                       [inv_num,name,location,status,comment,photo,item_id,uid()])
            flash(f"«{name}» обновлён.", "success")
        else:
            db.execute("INSERT INTO items(user_id,inv_num,name,location,status,comment,photo)"
                       " VALUES(?,?,?,?,?,?,?)", [uid(),inv_num,name,location,status,comment,photo])
            flash(f"«{name}» добавлен.", "success")
        db.commit()
        return redirect(url_for("index"))
    return render_template("item_form.html", statuses=ITEM_STATUSES, item=item)

@app.route("/items/delete/<int:item_id>", methods=["POST"])
@login_required
def item_delete(item_id):
    db   = get_db()
    item = db.execute("SELECT * FROM items WHERE id=? AND user_id=?",
                      [item_id, uid()]).fetchone()
    if item:
        if item["photo"]:
            p = os.path.join(UPLOAD_FOLDER, item["photo"])
            if os.path.exists(p): os.remove(p)
        db.execute("DELETE FROM items WHERE id=? AND user_id=?", [item_id, uid()])
        db.commit()
        flash(f"«{item['name']}» удалён.", "success")
    return redirect(url_for("index"))

# ════════════════════════════════════════════════════════════
#  НОМЕНКЛАТУРА
# ════════════════════════════════════════════════════════════
@app.route("/nomenclature")
@login_required
def nom_list():
    db = get_db()
    search = request.args.get("q","").strip()
    low_f  = request.args.get("low","")
    q = "SELECT * FROM nomenclature WHERE user_id=?"
    p = [uid()]
    if search:
        q += " AND (code LIKE ? OR name LIKE ? OR description LIKE ?)"
        like = f"%{search}%"; p += [like]*3
    if low_f: q += " AND min_qty>0 AND quantity<=min_qty"
    q += " ORDER BY name"
    items = db.execute(q, p).fetchall()
    return render_template("nom_list.html", items=items, search=search, low_f=low_f, units=UNITS)

@app.route("/nomenclature/add", methods=["GET","POST"])
@app.route("/nomenclature/edit/<int:nom_id>", methods=["GET","POST"])
@login_required
def nom_form(nom_id=None):
    db   = get_db()
    item = db.execute("SELECT * FROM nomenclature WHERE id=? AND user_id=?",
                      [nom_id, uid()]).fetchone() if nom_id else None
    if request.method == "POST":
        code  = request.form.get("code","").strip() or None
        name  = request.form["name"].strip()
        unit  = request.form.get("unit","шт")
        min_q = float(request.form.get("min_qty",0) or 0)
        desc  = request.form.get("description","").strip()
        if not name:
            flash("Название обязательно!", "error")
            return render_template("nom_form.html", item=item, units=UNITS)
        if nom_id:
            db.execute("UPDATE nomenclature SET code=?,name=?,unit=?,min_qty=?,description=?"
                       " WHERE id=? AND user_id=?", [code,name,unit,min_q,desc,nom_id,uid()])
            flash(f"«{name}» обновлён.", "success")
        else:
            iq = float(request.form.get("init_qty",0) or 0)
            db.execute("INSERT INTO nomenclature(user_id,code,name,unit,quantity,min_qty,description)"
                       " VALUES(?,?,?,?,?,?,?)", [uid(),code,name,unit,iq,min_q,desc])
            new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            if iq > 0:
                db.execute("INSERT INTO movements(user_id,doc_num,move_type,nom_id,quantity,comment)"
                           " VALUES(?,?,?,?,?,?)",
                           [uid(), next_doc(db,uid(),"НАЧ"), "receipt", new_id, iq, "Начальный остаток"])
            flash(f"«{name}» добавлен.", "success")
        db.commit()
        return redirect(url_for("nom_list"))
    return render_template("nom_form.html", item=item, units=UNITS)

@app.route("/nomenclature/delete/<int:nom_id>", methods=["POST"])
@login_required
def nom_delete(nom_id):
    db   = get_db()
    item = db.execute("SELECT * FROM nomenclature WHERE id=? AND user_id=?",
                      [nom_id, uid()]).fetchone()
    if item:
        if db.execute("SELECT 1 FROM movements WHERE nom_id=? AND user_id=?",
                      [nom_id, uid()]).fetchone():
            flash("Нельзя удалить: есть документы движения.", "error")
            return redirect(url_for("nom_list"))
        db.execute("DELETE FROM nomenclature WHERE id=? AND user_id=?", [nom_id, uid()])
        db.commit()
        flash(f"«{item['name']}» удалён.", "success")
    return redirect(url_for("nom_list"))

@app.route("/nomenclature/<int:nom_id>")
@login_required
def nom_detail(nom_id):
    db   = get_db()
    item = db.execute("SELECT * FROM nomenclature WHERE id=? AND user_id=?",
                      [nom_id, uid()]).fetchone()
    if not item:
        flash("Позиция не найдена.", "error"); return redirect(url_for("nom_list"))
    moves = db.execute("SELECT * FROM movements WHERE nom_id=? AND user_id=?"
                       " ORDER BY created_at DESC LIMIT 50", [nom_id, uid()]).fetchall()
    return render_template("nom_detail.html", item=item, moves=moves, move_types=MOVE_TYPES)

# ════════════════════════════════════════════════════════════
#  ДВИЖЕНИЯ
# ════════════════════════════════════════════════════════════
@app.route("/movements")
@login_required
def mov_list():
    db = get_db()
    search    = request.args.get("q","").strip()
    type_f    = request.args.get("type","")
    date_from = request.args.get("date_from","")
    date_to   = request.args.get("date_to","")
    q = ("SELECT m.*,n.name as nom_name,n.unit as nom_unit FROM movements m"
         " JOIN nomenclature n ON n.id=m.nom_id WHERE m.user_id=?")
    p = [uid()]
    if search:
        q += " AND (m.doc_num LIKE ? OR n.name LIKE ? OR m.comment LIKE ?)"
        like = f"%{search}%"; p += [like]*3
    if type_f and type_f in MOVE_TYPES:
        q += " AND m.move_type=?"; p.append(type_f)
    if date_from: q += " AND DATE(m.created_at)>=?"; p.append(date_from)
    if date_to:   q += " AND DATE(m.created_at)<=?"; p.append(date_to)
    q += " ORDER BY m.created_at DESC LIMIT 300"
    moves = db.execute(q, p).fetchall()
    return render_template("mov_list.html", moves=moves, move_types=MOVE_TYPES,
                           search=search, type_f=type_f, date_from=date_from, date_to=date_to)

@app.route("/movements/add", methods=["GET","POST"])
@login_required
def mov_add():
    db        = get_db()
    nom_items = db.execute("SELECT * FROM nomenclature WHERE user_id=? ORDER BY name",[uid()]).fetchall()
    preselect = request.args.get("nom_id","")
    if request.method == "POST":
        nom_id    = int(request.form["nom_id"])
        move_type = request.form["move_type"]
        qty       = float(request.form["quantity"])
        comment   = request.form.get("comment","").strip()
        if qty <= 0:
            flash("Количество > 0!", "error")
            return render_template("mov_form.html", nom_items=nom_items,
                                   move_types=MOVE_TYPES, preselect=preselect)
        nom = db.execute("SELECT * FROM nomenclature WHERE id=? AND user_id=?",
                         [nom_id, uid()]).fetchone()
        if not nom:
            flash("Позиция не найдена.", "error"); return redirect(url_for("mov_list"))
        if move_type in ("expense","writeoff") and nom["quantity"] < qty:
            flash(f"Недостаточно! Остаток: {nom['quantity']} {nom['unit']}", "error")
            return render_template("mov_form.html", nom_items=nom_items,
                                   move_types=MOVE_TYPES, preselect=preselect)
        doc = next_doc(db, uid())
        db.execute("INSERT INTO movements(user_id,doc_num,move_type,nom_id,quantity,comment)"
                   " VALUES(?,?,?,?,?,?)", [uid(), doc, move_type, nom_id, qty, comment])
        apply_mov(db, nom_id, move_type, qty)
        db.commit()
        flash(f"Документ {doc} проведён.", "success")
        return redirect(url_for("mov_list"))
    return render_template("mov_form.html", nom_items=nom_items,
                           move_types=MOVE_TYPES, preselect=preselect)

@app.route("/movements/delete/<int:mov_id>", methods=["POST"])
@login_required
def mov_delete(mov_id):
    db = get_db()
    m  = db.execute("SELECT * FROM movements WHERE id=? AND user_id=?",
                    [mov_id, uid()]).fetchone()
    if m:
        rev = {"receipt":"expense","expense":"receipt","writeoff":"receipt","return":"expense"}
        apply_mov(db, m["nom_id"], rev.get(m["move_type"],"expense"), m["quantity"])
        db.execute("DELETE FROM movements WHERE id=? AND user_id=?", [mov_id, uid()])
        db.commit()
        flash(f"Документ {m['doc_num']} отменён.", "success")
    return redirect(url_for("mov_list"))

# ════════════════════════════════════════════════════════════
#  ОТЧЁТЫ
# ════════════════════════════════════════════════════════════
@app.route("/reports")
@login_required
def reports():
    db = get_db()
    date_from = request.args.get("date_from", datetime.now().strftime("%Y-%m-01"))
    date_to   = request.args.get("date_to",   datetime.now().strftime("%Y-%m-%d"))
    summary = db.execute("""
        SELECT n.id, n.name, n.unit,
            SUM(CASE WHEN m.move_type='receipt'  THEN m.quantity ELSE 0 END) as total_receipt,
            SUM(CASE WHEN m.move_type='expense'  THEN m.quantity ELSE 0 END) as total_expense,
            SUM(CASE WHEN m.move_type='writeoff' THEN m.quantity ELSE 0 END) as total_writeoff,
            SUM(CASE WHEN m.move_type='return'   THEN m.quantity ELSE 0 END) as total_return,
            n.quantity as current_qty, n.min_qty
        FROM nomenclature n
        LEFT JOIN movements m ON m.nom_id=n.id AND m.user_id=n.user_id
            AND DATE(m.created_at) BETWEEN ? AND ?
        WHERE n.user_id=?
        GROUP BY n.id ORDER BY n.name
    """, [date_from, date_to, uid()]).fetchall()
    return render_template("reports.html", summary=summary,
                           date_from=date_from, date_to=date_to)

@app.route("/reports/export")
@login_required
def report_export():
    db = get_db()
    date_from = request.args.get("date_from", datetime.now().strftime("%Y-%m-01"))
    date_to   = request.args.get("date_to",   datetime.now().strftime("%Y-%m-%d"))
    summary = db.execute("""
        SELECT n.code,n.name,n.unit,
            SUM(CASE WHEN m.move_type='receipt'  THEN m.quantity ELSE 0 END) as r,
            SUM(CASE WHEN m.move_type='expense'  THEN m.quantity ELSE 0 END) as e,
            SUM(CASE WHEN m.move_type='writeoff' THEN m.quantity ELSE 0 END) as w,
            SUM(CASE WHEN m.move_type='return'   THEN m.quantity ELSE 0 END) as ret,
            n.quantity as qty, n.min_qty
        FROM nomenclature n
        LEFT JOIN movements m ON m.nom_id=n.id AND m.user_id=n.user_id
            AND DATE(m.created_at) BETWEEN ? AND ?
        WHERE n.user_id=? GROUP BY n.id ORDER BY n.name
    """, [date_from, date_to, uid()]).fetchall()
    movements = db.execute("""
        SELECT m.doc_num,m.move_type,m.created_at,n.name,n.unit,m.quantity,m.comment
        FROM movements m JOIN nomenclature n ON n.id=m.nom_id
        WHERE m.user_id=? AND DATE(m.created_at) BETWEEN ? AND ?
        ORDER BY m.created_at DESC
    """, [uid(), date_from, date_to]).fetchall()
    items = db.execute("SELECT * FROM items WHERE user_id=? ORDER BY updated_at DESC", [uid()]).fetchall()

    wb = openpyxl.Workbook()
    hf = PatternFill("solid", fgColor="1a1d27")
    hfont = Font(bold=True, color="FFFFFF")
    def hdr(ws, cols):
        for c,h in enumerate(cols,1):
            cell = ws.cell(row=1,column=c,value=h)
            cell.fill = hf; cell.font = hfont
            cell.alignment = Alignment(horizontal="center")

    ws1 = wb.active; ws1.title="Остатки"
    hdr(ws1,["Код","Наименование","Ед.","Приход","Расход","Списание","Возврат","Остаток","Мин."])
    for r,row in enumerate(summary,2):
        low = row["min_qty"]>0 and row["qty"]<=row["min_qty"]
        for c,v in enumerate([row["code"] or "",row["name"],row["unit"],
                               row["r"],row["e"],row["w"],row["ret"],row["qty"],row["min_qty"]],1):
            cell = ws1.cell(row=r,column=c,value=v)
            if low: cell.fill=PatternFill("solid",fgColor="2d0a0a"); cell.font=Font(color="f87171")
    for col,w in zip("ABCDEFGHI",[12,30,6,10,10,10,10,10,8]):
        ws1.column_dimensions[col].width=w

    ws2 = wb.create_sheet("Движения")
    hdr(ws2,["Документ","Тип","Дата","Наименование","Ед.","Количество","Комментарий"])
    tclr={"receipt":"052e16","expense":"2c1200","writeoff":"2d0a0a","return":"1e0a3c"}
    tn={k:v[0] for k,v in MOVE_TYPES.items()}
    for r,m in enumerate(movements,2):
        for c,v in enumerate([m["doc_num"],tn.get(m["move_type"],m["move_type"]),
                               m["created_at"][:16],m["name"],m["unit"],m["quantity"],m["comment"] or ""],1):
            ws2.cell(row=r,column=c,value=v).fill=PatternFill("solid",fgColor=tclr.get(m["move_type"],"111111"))
    for col,w in zip("ABCDEFG",[14,14,16,30,6,12,35]):
        ws2.column_dimensions[col].width=w

    ws3 = wb.create_sheet("Имущество")
    hdr(ws3,["Инв.номер","Наименование","Местонахождение","Статус","Комментарий","Обновлён"])
    sclr={"active":"052e16","broken":"2d0a0a","repair":"2a1f00","writeoff":"1e1e20","reserved":"1e0a3c"}
    for r,it in enumerate(items,2):
        for c,v in enumerate([it["inv_num"],it["name"],it["location"] or "",
                ITEM_STATUSES.get(it["status"],it["status"]),it["comment"] or "",it["updated_at"][:16]],1):
            ws3.cell(row=r,column=c,value=v).fill=PatternFill("solid",fgColor=sclr.get(it["status"],"111111"))
    for col,w in zip("ABCDEF",[14,28,20,18,35,16]):
        ws3.column_dimensions[col].width=w

    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

@app.route("/export/items")
@login_required
def export_items():
    db = get_db()
    items = db.execute("SELECT * FROM items WHERE user_id=? ORDER BY updated_at DESC",[uid()]).fetchall()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Имущество"
    hf=PatternFill("solid",fgColor="1a1d27"); hfont=Font(bold=True,color="FFFFFF")
    for c,h in enumerate(["Инв.номер","Наименование","Местонахождение","Статус","Комментарий","Обновлён"],1):
        cell=ws.cell(row=1,column=c,value=h); cell.fill=hf; cell.font=hfont
    sclr={"active":"d1fae5","broken":"fee2e2","repair":"fef9c3","writeoff":"f3f4f6","reserved":"ede9fe"}
    for r,it in enumerate(items,2):
        for c,v in enumerate([it["inv_num"],it["name"],it["location"] or "",
                ITEM_STATUSES.get(it["status"],it["status"]),it["comment"] or "",it["updated_at"][:16]],1):
            ws.cell(row=r,column=c,value=v).fill=PatternFill("solid",fgColor=sclr.get(it["status"],"FFFFFF"))
    for col,w in zip("ABCDEF",[14,28,20,18,35,16]):
        ws.column_dimensions[col].width=w
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"items_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

if __name__ == "__main__":
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    init_db()
    app.run(debug=True, host="0.0.0.0", port=5000)
